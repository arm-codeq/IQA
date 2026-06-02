import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
from copy import copy

# === การกำหนดค่าเริ่มต้น ===
folder = Path("IQA")
output_folder = Path("FN_sort_IQA") 
output_folder.mkdir(parents=True, exist_ok=True) 

target_files = ['GOV.xlsx', 'PRIV.xlsx', 'OTC.xlsx']

# 💡 กำหนดชื่อสารที่ต้องการ (ใส่ '' หากต้องการดึงยาทุกตัว)
TARGET_MOLECULE = 'MESALAZINE'

if TARGET_MOLECULE:
    output_filename = f"{TARGET_MOLECULE}_Volume_Sort_IQA.xlsx"
else:
    output_filename = "Combined_Volume_Sort_IQA.xlsx"

output_file_path = output_folder / output_filename

sort_col = 'Values (WAP) 2024' 
GROUP_COL_1 = 'Pack Molecule String'
GROUP_COL_2 = 'Manufacturer'
DETAIL_COL = 'Pack'

# === 🌟 1. ฟังก์ชันดึง "ความแรง + รูปแบบยา" เพื่อแยกประเภทชีต 🌟 ===
def get_dose_and_form(pack_name):
    if not isinstance(pack_name, str): return "Others"
    text = " ".join(pack_name.split()).upper()
    
    # 1. ดึงความแรง (Dose) เหมือนเดิม
    dose_match = re.search(r'(\d+(?:\.\d+)?\s*(?:MG|G|ML|L|MCG|IU|%)(?:\s*(?:/|-)?\s*\d+(?:\.\d+)?\s*(?:MG|G|ML|L|MCG|IU|%))*)', text)
    dose = dose_match.group(1).strip() if dose_match else ""
    
    # 2. เงื่อนไขดักจับรูปแบบยา (Form) จากข้อมูลทั้งหมดในเซ็ต
    form = ""
    if re.search(r'\b(TAB|TB|TABLET|FILM-COAT)\b', text):
        form = "TAB"
    elif re.search(r'\b(CAP|CAPSULE|SOFT)\b', text):
        form = "CAP"
    elif re.search(r'\b(SUPPOS|SUPP|SUPPOSITORIES)\b', text):
        form = "SUPPOS"
    elif re.search(r'\b(GRAN|GRANS|GRANULES)\b', text):
        form = "GRAN"
    elif re.search(r'\bENEMA\b', text):
        form = "ENEMA"
    elif re.search(r'\b(CRM|CREAM)\b', text):
        form = "CREAM"
    elif re.search(r'\b(OINT|OINTMENT)\b', text):
        form = "OINTMENT"
    elif re.search(r'\bBALM\b', text):
        form = "BALM"
    elif re.search(r'\bPASTE\b', text):
        form = "PASTE"
    elif re.search(r'\b(SYR|SYRUP)\b', text):
        form = "SYRUP"
    elif re.search(r'\b(SUSP|SUSPENSION)\b', text):
        form = "SUSPENSION"
    elif re.search(r'\b(SOL|SOLN|SOLUTION)\b', text):
        form = "SOLUTION"
    elif re.search(r'\b(DRP|DROP|DROPS)\b', text):
        form = "DROPS"
    elif re.search(r'\b(MXT|MIXTURE)\b', text):
        form = "MIXTURE"
    elif re.search(r'\bGEL\b', text):
        form = "GEL"
    elif re.search(r'\b(LOT|LOTION)\b', text):
        form = "LOTION"
    elif re.search(r'\b(SHAMPOO|SHPO)\b', text):
        form = "SHAMPOO"
    elif re.search(r'\b(PWD|POWDER)\b', text):
        form = "POWDER"
    elif re.search(r'\b(SACHET|SAC)\b', text):
        form = "SACHET"
    elif re.search(r'\b(VIAL|AMP|AMP\.|AMPOULE|PREFILL|SYRG)\b', text):
        form = "INJECTION"
    elif re.search(r'\bSPRAY\b', text):
        form = "SPRAY"
    elif re.search(r'\b(INHA|INHALER)\b', text):
        form = "INHALER"
    elif re.search(r'\b(PLASTER|PLAST|PATCH)\b', text):
        form = "PLASTER"
    elif re.search(r'\b(LOZ|LOZENGE)\b', text):
        form = "LOZENGE"
    elif re.search(r'\bOIL\b', text):
        form = "OIL"
    elif re.search(r'\b(INF|INFUSION)\b', text):
        form = "INFUSION"
    elif re.search(r'\bSOAP\b', text):
        form = "SOAP"
    elif re.search(r'\bTOOTHPASTE\b', text):
        form = "TOOTHPASTE"
    else:
        form = ""

    # 3. ประกอบชื่อสำหรับใช้ตั้งชื่อกลุ่มชีต
    if dose and form:
        return f"{dose} {form}"
    elif dose:
        return dose
    return "Others"

# === REFINED MULTIPLIER LOGIC ===
def get_pack_multiplier(pack_name):
    if not isinstance(pack_name, str): return 1.0
    text = " ".join(pack_name.split()).upper()
    if re.search(r'\bY\b', text): return 1.0
    num_pattern = r'(\d*\.?\d+)'
    unit_pattern = r'(?:ML|G|L|VIA|AMP|MG|SAC|SACHET)'

    if "SPRAY" in text:
        match_x_unit = re.search(f'{num_pattern}\s*[xX]\s*{num_pattern}\s*{unit_pattern}', text)
        if match_x_unit and "NASAL" not in text: return float(match_x_unit.group(1))
        return 1.0

    match_x_unit = re.search(f'{num_pattern}\s*[xX]\s*{num_pattern}\s*{unit_pattern}', text)
    if match_x_unit: return float(match_x_unit.group(1))

    match_x = re.search(f'{num_pattern}\s*[xX]\s*{num_pattern}$', text)
    if match_x: return float(match_x.group(1)) * float(match_x.group(2))

    match_space = re.search(f'{num_pattern}\s+{num_pattern}$', text)
    if match_space: return float(match_space.group(1)) * float(match_space.group(2))

    if re.search(f'{num_pattern}\s*{unit_pattern}$', text): return 1.0 

    match_end = re.search(f'{num_pattern}$', text)
    if match_end: return float(match_end.group(1))
    return 1.0

print(f"กำลังเตรียมสร้างไฟล์: {output_file_path.name}")

# === STEP 1: โหลดและจัดเตรียมข้อมูลดิบทั้งหมด ===
raw_data_dict = {}
master_cols = []

for file_name in target_files:
    file_path = folder / file_name
    print(f"📁 กำลังดึงข้อมูลจากไฟล์: {file_name}")
    
    try:
        sheets_dict = pd.read_excel(file_path, sheet_name=None, keep_default_na=False, na_values=[''])
    except FileNotFoundError:
        print(f"❌ 不พบไฟล์: {file_path}")
        continue

    prefix = Path(file_name).stem.upper()
    combined_file_df = []
    
    for sheet_name, df in sheets_dict.items():
        required_cols = [GROUP_COL_1, GROUP_COL_2]
        if not all(col in df.columns for col in required_cols): continue
        
        if not master_cols: master_cols = df.columns.tolist()

        num_cols = [c for c in df.columns if 'Values' in c or 'Units' in c]
        for col in num_cols:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')

        for c in required_cols + [DETAIL_COL]:
            if c in df.columns:
                df[c] = df[c].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
                
        df[GROUP_COL_1] = df[GROUP_COL_1].replace({'nan': np.nan, '': np.nan}).ffill()
        df[GROUP_COL_2] = df[GROUP_COL_2].replace({'nan': np.nan, '': np.nan}).ffill()

        is_subtotal = df[GROUP_COL_2].str.contains('Subtotal', na=False) | df[GROUP_COL_1].str.contains('Subtotal', na=False)
        data_df = df[~is_subtotal].copy()
        
        if TARGET_MOLECULE:
            data_df = data_df[data_df[GROUP_COL_1].str.upper() == TARGET_MOLECULE.upper()]
        
        if data_df.empty: continue
            
        data_df['__PackSize'] = data_df[DETAIL_COL].apply(get_pack_multiplier)
        # 🌟 เรียกใช้ฟังก์ชันเงื่อนไขรวมแบบใหม่ตรงนี้
        data_df['Dose_Label'] = data_df[DETAIL_COL].apply(get_dose_and_form)
        
        combined_file_df.append(data_df)
        
    if combined_file_df:
        raw_data_dict[prefix] = pd.concat(combined_file_df, ignore_index=True)

all_doses = set()
for df in raw_data_dict.values():
    all_doses.update(df['Dose_Label'].unique())

# === ฟังก์ชันสำหรับคำนวณ Block ข้อมูลย่อย ===
def create_block(data_df, cols_template):
    if data_df is None or data_df.empty: return pd.DataFrame(), pd.DataFrame(), []
    
    final_cols = cols_template.copy()
    for yr in ['2023', '2024']:
        unit_col = next((c for c in cols_template if 'Units' in c and yr in c and 'Growth' not in c), None)
        if unit_col:
            idx = final_cols.index(unit_col) + 1
            new_col = f"Per Tablet {yr}"
            if new_col not in final_cols: final_cols.insert(idx, new_col)

    unit_23_col = next((c for c in final_cols if 'Units' in c and '2023' in c and 'Growth' not in c), None)
    unit_24_col = next((c for c in final_cols if 'Units' in c and '2024' in c and 'Growth' not in c), None)
    
    if unit_23_col and unit_23_col in data_df: data_df[unit_23_col] = data_df[unit_23_col].fillna(0) * data_df['__PackSize']
    if unit_24_col and unit_24_col in data_df: data_df[unit_24_col] = data_df[unit_24_col].fillna(0) * data_df['__PackSize']

    calc_cols = [c for c in final_cols if ('Values' in c or 'Units' in c) and c in data_df.columns] 
    
    val_23_col = next((c for c in final_cols if 'Values' in c and '2023' in c), None)
    val_24_col = next((c for c in final_cols if 'Values' in c and '2024' in c), None)
    val_growth_col = next((c for c in final_cols if 'Values' in c and 'Growth' in c), None)
    unit_growth_col = next((c for c in final_cols if 'Units' in c and 'Growth' in c), None)

    def recalculate_growth(df_target):
        if val_23_col and val_24_col and val_growth_col and val_23_col in df_target and val_24_col in df_target:
            df_target[val_growth_col] = (df_target[val_24_col] - df_target[val_23_col]) / df_target[val_23_col].replace(0, np.nan)
        if unit_23_col and unit_24_col and unit_growth_col and unit_23_col in df_target and unit_24_col in df_target:
            df_target[unit_growth_col] = (df_target[unit_24_col] - df_target[unit_23_col]) / df_target[unit_23_col].replace(0, np.nan)
        return df_target

    manuf_totals = data_df.groupby([GROUP_COL_1, GROUP_COL_2], dropna=False)[calc_cols].sum().reset_index()
    manuf_totals = recalculate_growth(manuf_totals)
    pack_totals = manuf_totals.groupby(GROUP_COL_1, dropna=False)[calc_cols].sum().reset_index()
    pack_totals = recalculate_growth(pack_totals)

    pack_order = pack_totals.sort_values(sort_col, ascending=False)[GROUP_COL_1].tolist() if sort_col in pack_totals.columns else pack_totals[GROUP_COL_1].tolist()

    out_rows = []
    for pack_name in pack_order:
        pack_data = data_df[data_df[GROUP_COL_1] == pack_name].copy()
        manufs_in_pack = manuf_totals[manuf_totals[GROUP_COL_1] == pack_name].sort_values(sort_col, ascending=False)[GROUP_COL_2].tolist() if sort_col in manuf_totals.columns else manuf_totals[manuf_totals[GROUP_COL_1] == pack_name][GROUP_COL_2].tolist()
        top_manuf_for_pack = manufs_in_pack[0] if manufs_in_pack else None

        for manuf_name in manufs_in_pack:
            grp = pack_data[pack_data[GROUP_COL_2] == manuf_name]
            if sort_col in grp.columns: grp = grp.sort_values(sort_col, ascending=False)
            top_idx = grp.index[0]
            for i, r in grp.iterrows():
                row = {c: np.nan for c in final_cols}
                for c in cols_template:
                    if c in r: row[c] = r[c]
                row[GROUP_COL_2] = manuf_name if i == top_idx else ""
                row[GROUP_COL_1] = pack_name if (top_manuf_for_pack == manuf_name and i == top_idx) else ""
                out_rows.append(row)

            manuf_sub_row = {c: np.nan for c in final_cols}
            manuf_sub_row[GROUP_COL_2] = f"{manuf_name} Subtotal"
            manuf_total_data = manuf_totals[(manuf_totals[GROUP_COL_1] == pack_name) & (manuf_totals[GROUP_COL_2] == manuf_name)].iloc[0]
            for col in calc_cols: manuf_sub_row[col] = manuf_total_data[col]
            out_rows.append(manuf_sub_row)

        pack_sub_row = {c: np.nan for c in final_cols}
        pack_sub_row[GROUP_COL_1] = f"{pack_name} Subtotal"
        pack_total_data = pack_totals[pack_totals[GROUP_COL_1] == pack_name].iloc[0]
        for col in calc_cols: pack_sub_row[col] = pack_total_data[col]
        out_rows.append(pack_sub_row)

    final_df = pd.DataFrame(out_rows, columns=final_cols)

    def insert_per_tablet(df_insert, year_suffix):
        val_col = next((c for c in df_insert.columns if 'Values' in c and year_suffix in c), None)
        unit_col = next((c for c in df_insert.columns if 'Units' in c and year_suffix in c and 'Growth' not in c), None)
        per_tab_col = f"Per Tablet {year_suffix}"
        if val_col and unit_col and per_tab_col in df_insert.columns:
            vals = pd.to_numeric(df_insert[val_col], errors='coerce').fillna(0)
            units_tablets = pd.to_numeric(df_insert[unit_col], errors='coerce').fillna(0)
            per_tab_values = vals / units_tablets.replace(0, np.nan) 
            df_insert[per_tab_col] = per_tab_values
            mask_sub = (df_insert[GROUP_COL_1].astype(str).str.contains('Subtotal', na=False) | df_insert[GROUP_COL_2].astype(str).str.contains('Subtotal', na=False))
            df_insert.loc[mask_sub, per_tab_col] = np.nan

    insert_per_tablet(final_df, '2023')
    insert_per_tablet(final_df, '2024')

    for c in final_cols:
        if 'Values' in c or 'Units' in c or 'Per Tablet' in c:
            final_df[c] = final_df[c].map(lambda x: "-" if pd.isna(x) or x == 0 else x)

    return final_df, data_df, final_cols

def make_summary_row(raw_dfs_list, label_name, final_cols):
    if not raw_dfs_list: return None
    combined_raw = pd.concat(raw_dfs_list, ignore_index=True)
    row_dict = {c: np.nan for c in final_cols}
    row_dict[GROUP_COL_1] = label_name
    
    v23_col = next((c for c in final_cols if 'Values' in c and '2023' in c), None)
    v24_col = next((c for c in final_cols if 'Values' in c and '2024' in c), None)
    u23_col = next((c for c in final_cols if 'Units' in c and '2023' in c and 'Growth' not in c), None)
    u24_col = next((c for c in final_cols if 'Units' in c and '2024' in c and 'Growth' not in c), None)
    vg_col = next((c for c in final_cols if 'Values' in c and 'Growth' in c), None)
    ug_col = next((c for c in final_cols if 'Units' in c and 'Growth' in c), None)
    pt23_col = next((c for c in final_cols if 'Per Tablet' in c and '2023' in c), None)
    pt24_col = next((c for c in final_cols if 'Per Tablet' in c and '2024' in c), None)
    
    for c in [v23_col, v24_col, u23_col, u24_col]:
        if c and c in combined_raw.columns: row_dict[c] = combined_raw[c].sum()
        
    v23 = row_dict.get(v23_col, 0) or 0
    v24 = row_dict.get(v24_col, 0) or 0
    if vg_col and v23 != 0: row_dict[vg_col] = (v24 - v23) / v23
    
    u23 = row_dict.get(u23_col, 0) or 0
    u24 = row_dict.get(u24_col, 0) or 0
    if ug_col and u23 != 0: row_dict[ug_col] = (u24 - u23) / u23
    
    if "GOV+PRIV Subtotal" not in str(label_name) and "Grand Total" not in str(label_name):
        if pt23_col and u23 != 0: row_dict[pt23_col] = v23 / u23
        if pt24_col and u24 != 0: row_dict[pt24_col] = v24 / u24
    
    for c in final_cols:
        if c not in [GROUP_COL_1, GROUP_COL_2]:
            val = row_dict.get(c)
            if pd.isna(val) or val == 0: row_dict[c] = "-"
    return pd.DataFrame([row_dict])

# === STEP 2: ประมวลผลและสร้าง Sheet สำหรับแต่ละปริมาณ (Dose) ===
sheet_layout_info = {}

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    for dose in sorted(list(all_doses)):
        dose_safe_name = re.sub(r'[\\/*?:\[\]]', '_', dose)[:31]
        
        gov_raw = raw_data_dict.get('GOV')
        gov_raw = gov_raw[gov_raw['Dose_Label'] == dose] if gov_raw is not None else None
        priv_raw = raw_data_dict.get('PRIV')
        priv_raw = priv_raw[priv_raw['Dose_Label'] == dose] if priv_raw is not None else None
        otc_raw = raw_data_dict.get('OTC')
        otc_raw = otc_raw[otc_raw['Dose_Label'] == dose] if otc_raw is not None else None

        gov_final, gov_raw_processed, final_cols = create_block(gov_raw, master_cols)
        priv_final, priv_raw_processed, _ = create_block(priv_raw, master_cols)
        otc_final, otc_raw_processed, _ = create_block(otc_raw, master_cols)

        if not final_cols: continue 

        combined_df_list = []
        raw_gov_priv = []
        raw_all = []
        len_gov = len_priv = len_otc = has_subtotal = 0
        
        mol_label = f"{TARGET_MOLECULE} ({dose})" if TARGET_MOLECULE else f"Total ({dose})"

        if not gov_final.empty:
            combined_df_list.append(gov_final); raw_gov_priv.append(gov_raw_processed); raw_all.append(gov_raw_processed)
            len_gov = len(gov_final)
            
        if not priv_final.empty:
            combined_df_list.append(priv_final); raw_gov_priv.append(priv_raw_processed); raw_all.append(priv_raw_processed)
            len_priv = len(priv_final)
            
        if raw_gov_priv:
            sub_row = make_summary_row(raw_gov_priv, f"{mol_label} GOV+PRIV Subtotal", final_cols)
            if sub_row is not None: combined_df_list.append(sub_row); has_subtotal = 1
            
        if not otc_final.empty:
            combined_df_list.append(otc_final); raw_all.append(otc_raw_processed)
            len_otc = len(otc_final)
            
        if raw_all:
            grand_row = make_summary_row(raw_all, f"{mol_label} Grand Total (GOV+PRIV+OTC)", final_cols)
            if grand_row is not None: combined_df_list.append(grand_row)

        if combined_df_list:
            df_ruam = pd.concat(combined_df_list, ignore_index=True)
            df_ruam.to_excel(writer, sheet_name=dose_safe_name, index=False, startrow=2)
            sheet_layout_info[dose_safe_name] = {
                'len_gov': len_gov,
                'len_priv': len_priv,
                'len_otc': len_otc,
                'has_sub': has_subtotal,
                'dose_label': dose 
            }

# === STEP 3: Styles & Print Setup สำหรับทุกชีต ===
print("🎨 จัดรูปแบบ, แทรก Headers และตั้งค่าการ Print...")
wb = load_workbook(output_file_path)

fill_gray_header = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") 
fill_gray_sub = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
fill_blue_gov_priv = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") 
fill_pink = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid") 
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 

font_std = Font(name="Tahoma", size=8)
font_bold = Font(name="Tahoma", size=8, bold=True) 
font_red_bold = Font(name="Tahoma", size=11, bold=True, color="FF0000") 

thin_side = Side(style='thin', color='D3D3D3')
thick_side = Side(style='thick', color='000000') 
border_lr = Border(left=thin_side, right=thin_side)
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_thick = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side) 

fmt_currency = '"฿"#,##0_-'; fmt_currency_2 = '"฿"#,##0.00_-'; fmt_number = '#,##0_-'; fmt_percent = '0.0%' 

def copy_header_format(ws, source_rows, target_start_row):
    for r_idx, row_to_copy in enumerate(source_rows, start=0):
        target_row = target_start_row + r_idx
        ws.row_dimensions[target_row].height = ws.row_dimensions[row_to_copy[0].row].height
        for cell in row_to_copy:
            new_cell = ws.cell(row=target_row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font); new_cell.border = copy(cell.border); new_cell.fill = copy(cell.fill); new_cell.number_format = copy(cell.number_format); new_cell.alignment = copy(cell.alignment)
    ws.merge_cells(start_row=target_start_row, start_column=4, end_row=target_start_row, end_column=6)
    ws.merge_cells(start_row=target_start_row, start_column=7, end_row=target_start_row, end_column=9)
    ws.merge_cells(start_row=target_start_row+1, start_column=4, end_row=target_start_row+1, end_column=6)
    ws.merge_cells(start_row=target_start_row+1, start_column=7, end_row=target_start_row+1, end_column=9)

for sheet_name_wb in wb.sheetnames:
    ws = wb[sheet_name_wb]
    info = sheet_layout_info.get(sheet_name_wb, {})
    len_gov = info.get('len_gov', 0)
    len_priv = info.get('len_priv', 0)
    len_otc = info.get('len_otc', 0)
    has_sub = info.get('has_sub', 0)
    dose_lbl = info.get('dose_label', sheet_name_wb)
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4             
    ws.page_margins.left = 0; ws.page_margins.right = 0; ws.page_margins.top = 0; ws.page_margins.bottom = 0
    ws.page_margins.header = 0; ws.page_margins.footer = 0
    ws.print_options.horizontalCentered = True; ws.print_options.verticalCentered = True

    ws['D1'] = "Relative MAT"; ws['G1'] = "Relative MAT"; ws['D2'] = "MAT 2023/12"; ws['G2'] = "MAT 2024/12"
    ws.merge_cells('D1:F1'); ws.merge_cells('G1:I1'); ws.merge_cells('D2:F2'); ws.merge_cells('G2:I2')
    for cell in ws[1]: cell.fill = fill_gray_header; cell.font = font_std; cell.alignment = Alignment(horizontal='center'); cell.border = border_all
    for cell in ws[2]: 
        cell.fill = fill_white; cell.font = font_std; cell.alignment = Alignment(horizontal='center'); cell.border = border_all
        if cell.column == 3: cell.font = font_bold

    headers = []
    for cell in ws[3]:
        cell.fill = fill_gray_header; cell.font = font_std; cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = border_all
        headers.append(str(cell.value).strip() if cell.value else "")

    try:
        idx_group1 = headers.index(GROUP_COL_1)
        idx_group2 = headers.index(GROUP_COL_2)
        idx_detail = headers.index(DETAIL_COL) if DETAIL_COL in headers else -1
        cols_to_clear_text = [idx_group1, idx_group2, idx_detail]
    except ValueError:
        cols_to_clear_text = []

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        try:
            pack_val = row[idx_group1].value; manu_val = row[idx_group2].value
        except: pack_val = manu_val = None
        
        is_subtotal = False; is_grand_total = False; is_gov_priv_sub = False
        row_fill = fill_white; row_border = border_all
        bold_cols = []; data_start_idx = 0
        
        for idx, cell in enumerate(row):
            if cell.value is not None and str(cell.value).strip() != "": data_start_idx = idx; break
            
        str_pack = str(pack_val or ""); str_manu = str(manu_val or "")
        
        if "Subtotal" in str_manu:
            row_fill = fill_gray_sub; bold_cols.append(idx_group2); is_subtotal = True
            
        elif "Subtotal" in str_pack or "Grand Total" in str_pack:
            bold_cols.append(idx_group1); is_subtotal = True
            if "GOV+PRIV Subtotal" in str_pack:
                row_fill = fill_pink; is_gov_priv_sub = True 
            elif "Grand Total" in str_pack:
                row_fill = fill_yellow; row_border = border_thick; is_grand_total = True 
            else:
                r_idx = row[0].row
                if r_idx <= (3 + len_gov + len_priv + has_sub):
                    row_fill = fill_blue_gov_priv 
                else:
                    row_fill = fill_blue_gov_priv 
            
        for col_idx, cell in enumerate(row):
            col_name = headers[col_idx] if col_idx < len(headers) else ""
            
            if col_idx < data_start_idx:
                cell.fill = fill_white; cell.border = border_lr; cell.font = font_std
            else:
                cell.border = row_border if is_subtotal else border_all
                cell.fill = row_fill if is_subtotal else fill_white
                if is_grand_total: cell.font = font_red_bold
                else: cell.font = font_bold if col_idx in bold_cols else font_std
            
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
                if any(x in col_name for x in ['%', 'Growth', 'Share']): cell.number_format = fmt_percent
                elif 'Per Tablet' in col_name: cell.number_format = fmt_currency_2
                elif 'Values' in col_name: cell.number_format = fmt_currency
                else: cell.number_format = fmt_number

            if (is_grand_total or is_gov_priv_sub) and (col_idx in cols_to_clear_text):
                cell.value = ""

    ws.row_dimensions[3].height = 30 
    
    if len_gov > 0: ws['C2'] = f"GOV {dose_lbl}"; ws['C2'].font = font_bold
    elif len_priv > 0: ws['C2'] = f"PRIV {dose_lbl}"; ws['C2'].font = font_bold
    elif len_otc > 0: ws['C2'] = f"OTC {dose_lbl}"; ws['C2'].font = font_bold

    # 2. แทรกแถวหัวตาราง (Insert Rows) แบบ Bottom-Up ป้องกันแถวเลื่อนสลับตำแหน่ง
    idx_otc = 4 + len_gov + len_priv + has_sub
    idx_priv = 4 + len_gov
    source_header = [[ws.cell(row=r, column=c) for c in range(1, ws.max_column+1)] for r in [1, 2, 3]]
    
    insert_otc = len_otc > 0 and (len_gov > 0 or len_priv > 0)
    insert_priv = len_priv > 0 and len_gov > 0
    
    if insert_otc:
        ws.insert_rows(idx_otc, 3)
        copy_header_format(ws, source_header, idx_otc)
        ws.cell(row=idx_otc + 1, column=3, value=f"OTC {dose_lbl}").font = font_bold 
        
    if insert_priv:
        ws.insert_rows(idx_priv, 3)
        copy_header_format(ws, source_header, idx_priv)
        ws.cell(row=idx_priv + 1, column=3, value=f"PRIV {dose_lbl}").font = font_bold 

    # 3. ปรับขนาดหน้ากว้างคอลัมน์ (Auto-fit Width) โดยไม่นำเอาข้อความหัวตารางใหม่มาคิดคำนวณปนกับตัวเลข
    target_autofit_cols = [DETAIL_COL, 'Values (WAP) 2023', 'Units 2023', 'Values (WAP) 2024', 'Units.1 2024', 'Units 2024']
    current_headers = [str(ws.cell(row=3, column=c).value).strip() if ws.cell(row=3, column=c).value else "" for c in range(1, ws.max_column + 1)]

    for col_idx, col_name in enumerate(current_headers, 1):
        if col_name in target_autofit_cols or any(t in col_name for t in ['Values (WAP) 2023', 'Units 2023', 'Values (WAP) 2024', 'Units 2024']):
            max_len = len(str(col_name))
            for r in range(4, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=col_idx).value
                if cell_val is not None and str(cell_val).strip() != "":
                    # ข้ามแถวที่เป็นโครงสร้างหัวตารางแทรกใหม่เพื่อความแม่นยำของขนาดคอลัมน์ตัวเลข
                    if r in [idx_priv, idx_priv+1, idx_priv+2] and insert_priv: continue
                    if r in [idx_otc, idx_otc+1, idx_otc+2] and insert_otc: continue
                    
                    if isinstance(cell_val, (int, float)):
                        if 'Units' in col_name:
                            formatted_str = f"{cell_val:,.0f}"
                        else:
                            formatted_str = f"{cell_val:,.2f}"
                        max_len = max(max_len, len(formatted_str))
                    else:
                        max_len = max(max_len, len(str(cell_val)))
            
            padding = 4 if col_name == DETAIL_COL else 1.5
            adjusted_width = min(max_len + padding, 80)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

wb.save(output_file_path)
print(f"✨ เสร็จสมบูรณ์! ไฟล์ประมวลผลแยกประเภทยารองรับเงื่อนไขทั้งหมดเรียบร้อย: {output_file_path.name}")